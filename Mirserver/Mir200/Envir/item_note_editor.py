# -*- coding: utf-8 -*-
"""
Item.xls 备注修改工具 (UBB 版)
支持 UBB 富文本格式：[size=xx] [color=xx] [br] [b] [i] [u]
所有标签使用方括号 [...] 格式
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser
from openpyxl import load_workbook
import xlrd
from xlutils.copy import copy as xl_copy
import os
import sys
import logging
import re

# UBB 标签正则表达式 (匹配所有 [tag], [/tag], [tag=val])
UBB_TAG_RE = re.compile(r'\[/?[a-zA-Z]+[^]]*\]', re.IGNORECASE)

def strip_ubb_tags(text):
    """移除文本中所有的 UBB 标签"""
    return UBB_TAG_RE.sub('', text)

class Logger:
    def __init__(self, log_file=None):
        self.log_file = log_file or os.path.join(os.path.dirname(os.path.abspath(__file__)), 'item_editor.log')
        self.logger = logging.getLogger('ItemEditor')
        self.logger.setLevel(logging.DEBUG)
        fh = logging.FileHandler(self.log_file, encoding='utf-8')
        fh.setLevel(logging.DEBUG)
        ch = logging.StreamHandler()
        ch.setLevel(logging.INFO)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        fh.setFormatter(formatter)
        ch.setFormatter(formatter)
        self.logger.addHandler(fh)
        self.logger.addHandler(ch)
    
    def info(self, msg): self.logger.info(msg)
    def warning(self, msg): self.logger.warning(msg)
    def error(self, msg): self.logger.error(msg)
    def get_log_content(self):
        try:
            with open(self.log_file, 'r', encoding='utf-8') as f: return f.read()
        except Exception: return '暂无日志'

class ItemNoteEditor:
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title('Item.xls 备注修改工具 (UBB版)')
        self.root.geometry('1200x850')
        
        self.base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        self.file_path = None
        self.file_type = None
        self.xlrd_book = None
        self.wb = None
        self.items = []
        self.current_row = None
        self.preview_timer = None
        
        self.logger = Logger()
        self.logger.info('=' * 50)
        self.logger.info('Item备注修改工具 (UBB版) 启动')
        self._setup_ui()
        self._try_load_default()

    def _setup_ui(self):
        toolbar = tk.Frame(self.root, bg='#f0f0f0', height=40)
        toolbar.pack(fill=tk.X, padx=5, pady=5)
        
        self.btn_open = tk.Button(toolbar, text='打开文件', command=self.open_file, bg='#4CAF50', fg='white', font=('微软雅黑', 9))
        self.btn_open.pack(side=tk.LEFT, padx=5)
        
        self.btn_reload = tk.Button(toolbar, text='重新读取文件', command=self._reload_file, bg='#FF9800', fg='white', font=('微软雅黑', 9))
        self.btn_reload.pack(side=tk.LEFT, padx=5)
        
        self.btn_refresh = tk.Button(toolbar, text='读取备注', command=self._refresh_desc, bg='#2196F3', fg='white', font=('微软雅黑', 9))
        self.btn_refresh.pack(side=tk.LEFT, padx=5)
        
        self.lbl_file = tk.Label(toolbar, text='未选择文件', bg='#f0f0f0', fg='#666', font=('微软雅黑', 9))
        self.lbl_file.pack(side=tk.LEFT, padx=10)
        
        self.btn_save = tk.Button(toolbar, text='保存备注', command=self._save_desc, bg='#4CAF50', fg='white', font=('微软雅黑', 9), state=tk.DISABLED)
        self.btn_save.pack(side=tk.RIGHT, padx=5)
        
        self.btn_log = tk.Button(toolbar, text='查看日志', command=self._show_log, bg='#607D8B', fg='white', font=('微软雅黑', 9))
        self.btn_log.pack(side=tk.RIGHT, padx=5)
        
        main_frame = tk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 左侧：物品列表
        left_frame = tk.Frame(main_frame, width=300)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 5))
        left_frame.pack_propagate(False)
        
        search_frame = tk.Frame(left_frame)
        search_frame.pack(fill=tk.X, pady=(0, 5))
        tk.Label(search_frame, text='搜索:', font=('微软雅黑', 10)).pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self._filter_items)
        tk.Entry(search_frame, textvariable=self.search_var, font=('微软雅黑', 9)).pack(fill=tk.X, expand=True, padx=(5, 0))
        
        list_frame = tk.Frame(left_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        self.tree = ttk.Treeview(list_frame, columns=('ID', 'Name'), show='headings', height=30)
        self.tree.heading('ID', text='ID')
        self.tree.heading('Name', text='道具名称')
        self.tree.column('ID', width=60, minwidth=50)
        self.tree.column('Name', width=200, minwidth=150)
        
        tree_scroll = tk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.tree.bind('<<TreeviewSelect>>', self._on_item_select)
        
        # 右侧：编辑区域
        right_frame = tk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        self.right_container = tk.Frame(right_frame)
        self.right_container.pack(fill=tk.BOTH, expand=True)
        
        welcome_frame = tk.Frame(self.right_container)
        welcome_frame.pack(expand=True)
        
        tk.Label(welcome_frame, text='Item.xls 备注修改工具 (UBB版)', 
                font=('微软雅黑', 16, 'bold'), fg='#333').pack(pady=20)
        
        # 更新显示的路径
        tk.Label(welcome_frame, text='默认读取 ..\\..\\Mir200\\Envir\\data\\Item.xls', 
                font=('微软雅黑', 10), fg='#666').pack()
        
        tk.Label(welcome_frame, text='修改备注后直接保存回原文件', 
                font=('微软雅黑', 9), fg='#FF9800').pack(pady=10)
        tk.Label(welcome_frame, text='支持UBB富文本：[size=xx] [color=xx] [br] [b] [i] [u]', 
                font=('微软雅黑', 9), fg='#4CAF50').pack(pady=5)

    def _try_load_default(self):
        # 修改后的默认路径：相对于 exe 目录的 ../../Mir200/Envir/data/
        default_xls = os.path.join(self.base_dir, '..', '..', 'Mir200', 'Envir', 'data', 'Item.xls')
        default_xlsx = os.path.join(self.base_dir, '..', '..', 'Mir200', 'Envir', 'data', 'Item.xlsx')
        
        self.logger.info(f'查找默认文件: {default_xls}')
        if os.path.exists(default_xlsx): self._load_file(default_xlsx)
        elif os.path.exists(default_xls): self._load_file(default_xls)
        else: self.logger.warning('未找到默认文件')

    def open_file(self):
        file_path = filedialog.askopenfilename(
            title='选择Item文件',
            filetypes=[('Excel文件', '*.xls *.xlsx')],
            initialdir=self.base_dir
        )
        if file_path: self._load_file(file_path)

    def _reload_file(self):
        if not self.file_path: return messagebox.showwarning('警告', '请先打开文件')
        try:
            self._load_file(self.file_path)
            messagebox.showinfo('成功', '文件已重新读取')
        except Exception as e: messagebox.showerror('错误', str(e))

    def _refresh_desc(self):
        if not self.current_row: return messagebox.showwarning('警告', '请先选择道具')
        try:
            row = self.current_row['row']
            desc = self.xlrd_book.sheet_by_index(0).cell_value(row, 25) if self.file_type == '.xls' else self.wb.active.cell(row=row, column=26).value or ''
            self.current_row['desc'] = str(desc) if desc else ''
            
            display = self.current_row['desc'].replace('[br]', '\n')
            self.editor.delete('1.0', tk.END)
            self.editor.insert('1.0', display)
            
            self.original_text.config(state=tk.NORMAL)
            self.original_text.delete('1.0', tk.END)
            self.original_text.insert('1.0', self.current_row['desc'])
            self.original_text.config(state=tk.DISABLED)
            
            self._update_preview()
            messagebox.showinfo('成功', f'已重新读取：{self.current_row["name"]}')
        except Exception as e: messagebox.showerror('错误', str(e))

    def _load_file(self, file_path):
        try:
            self.file_path = file_path
            self.file_type = os.path.splitext(file_path)[1].lower()
            self.logger.info(f'加载: {file_path} ({self.file_type})')
            
            if self.file_type == '.xls': self._load_xls(file_path)
            elif self.file_type == '.xlsx': self._load_xlsx(file_path)
            else: return messagebox.showwarning('警告', '仅支持 .xls 或 .xlsx')
            
            self.lbl_file.config(text=f'{os.path.basename(file_path)} ({self.file_type}) - {len(self.items)}个物品')
            self.btn_save.config(state=tk.NORMAL)
            self.btn_refresh.config(state=tk.NORMAL)
            self.current_row = None
        except Exception as e: messagebox.showerror('错误', str(e))

    def _load_xls(self, fp):
        self.xlrd_book = xlrd.open_workbook(fp, formatting_info=False)
        sh = self.xlrd_book.sheet_by_index(0)
        hdr = self._find_header(sh)
        self.items = []
        for r in range(hdr+1, sh.nrows):
            iid = self._try_id(sh.cell_value(r, 0))
            if iid is None: continue
            name = sh.cell_value(r, 1)
            if not name: continue
            self.items.append({'row': r, 'id': iid, 'name': str(name), 'desc': str(sh.cell_value(r, 25) or '')})
        self._populate_tree()

    def _load_xlsx(self, fp):
        self.wb = load_workbook(fp)
        ws = self.wb.active
        hdr = self._find_header_ws(ws)
        self.items = []
        for r in range(hdr+1, ws.max_row+1):
            iid = self._try_id(ws.cell(row=r, column=1).value)
            if iid is None: continue
            name = ws.cell(row=r, column=2).value
            if not name: continue
            self.items.append({'row': r, 'id': iid, 'name': str(name), 'desc': str(ws.cell(row=r, column=26).value or '')})
        self._populate_tree()

    def _find_header(self, sh):
        for r in range(min(10, sh.nrows)):
            if any('id' in str(sh.cell_value(r,c)).lower() for c in range(min(3, sh.ncols))): return r
        return 0

    def _find_header_ws(self, ws):
        for r in range(1, min(11, ws.max_row+1)):
            if any('id' in str(ws.cell(row=r, column=c).value or '').lower() for c in range(1,4)): return r
        return 1

    def _try_id(self, v):
        try: return int(float(v))
        except: return None

    def _populate_tree(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        for i in self.items: self.tree.insert('', tk.END, values=(i['id'], i['name']))

    def _filter_items(self, *args):
        st = self.search_var.get().lower()
        for i in self.tree.get_children(): self.tree.delete(i)
        for i in self.items:
            if st in str(i['name']).lower() or st in str(i['id']).lower(): self.tree.insert('', tk.END, values=(i['id'], i['name']))

    def _on_item_select(self, event):
        sel = self.tree.selection()
        if not sel: return
        iid = int(self.tree.item(sel[0])['values'][0])
        for i in self.items:
            if i['id'] == iid: self.current_row = i; self._show_editor(i); return

    def _show_editor(self, item):
        for w in self.right_container.winfo_children(): w.destroy()
        tk.Label(self.right_container, text=f'{item["name"]} (ID: {item["id"]})', font=('微软雅黑', 12, 'bold')).pack(anchor=tk.W, pady=(0,5))
        tk.Label(self.right_container, text='原始备注（只读）：', font=('微软雅黑', 9, 'bold')).pack(anchor=tk.W, pady=(5,2))
        self.original_text = tk.Text(self.right_container, height=4, wrap=tk.WORD, font=('微软雅黑', 9), bg='#f5f5f5', state=tk.DISABLED)
        self.original_text.insert('1.0', item['desc'])
        self.original_text.config(state=tk.DISABLED)
        self.original_text.pack(fill=tk.X, padx=5, pady=2)
        
        split = tk.PanedWindow(self.right_container, orient=tk.HORIZONTAL)
        split.pack(fill=tk.BOTH, expand=True, padx=5, pady=2)
        
        pfr = tk.Frame(split)
        split.add(pfr, width=400)
        tk.Label(pfr, text='富文本预览:', font=('微软雅黑', 9, 'bold')).pack(anchor=tk.W)
        self.preview = tk.Text(pfr, wrap=tk.WORD, font=('微软雅黑', 10), bg='#e8f5e9', state=tk.DISABLED)
        ps = tk.Scrollbar(pfr, command=self.preview.yview)
        self.preview.config(yscrollcommand=ps.set)
        self.preview.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        ps.pack(side=tk.RIGHT, fill=tk.Y)
        
        efr = tk.Frame(split)
        split.add(efr, width=400)
        tk.Label(efr, text='编辑源码 (UBB):', font=('微软雅黑', 9, 'bold')).pack(anchor=tk.W)
        self.editor = tk.Text(efr, wrap=tk.WORD, font=('Consolas', 10), undo=True, height=8)
        self.editor.insert('1.0', item['desc'].replace('[br]', '\n'))
        es = tk.Scrollbar(efr, command=self.editor.yview)
        self.editor.config(yscrollcommand=es.set)
        self.editor.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        es.pack(side=tk.RIGHT, fill=tk.Y)
        self.editor.bind('<KeyRelease>', lambda e: self._schedule_preview())
        
        tool = tk.Frame(self.right_container)
        tool.pack(fill=tk.X, padx=5, pady=5)
        
        tk.Label(tool, text='字号:', font=('微软雅黑', 9)).pack(side=tk.LEFT)
        self.size_var = tk.StringVar(value='选择')
        ttk.Combobox(tool, textvariable=self.size_var, values=[str(i) for i in range(10,81,2)], width=6, state='readonly').pack(side=tk.LEFT, padx=5)
        tk.Button(tool, text='应用字号', command=self._apply_size, width=8).pack(side=tk.LEFT, padx=5)
        
        # 自定义颜色按钮
        tk.Button(tool, text='自定义颜色', command=self._apply_custom_color, width=12, bg='#FFEB3B').pack(side=tk.LEFT, padx=20)
        
        tk.Button(tool, text='刷新预览', command=self._update_preview, width=8).pack(side=tk.RIGHT, padx=5)
        
        tk.Label(self.right_container, text='UBB 格式：[size=50]大字[/size] [color=red]红色[/color] [br]换行', font=('微软雅黑', 8), fg='#666').pack(anchor=tk.W, padx=5)
        
        self._update_preview()

    def _schedule_preview(self):
        if self.preview_timer: self.root.after_cancel(self.preview_timer)
        self.preview_timer = self.root.after(300, self._update_preview)

    def _update_preview(self):
        if not hasattr(self, 'preview'): return
        raw_text = self.editor.get('1.0', tk.END)
        
        self.preview.config(state=tk.NORMAL)
        self.preview.delete('1.0', tk.END)
        for t in self.preview.tag_names():
            if t.startswith('u_'): self.preview.tag_delete(t)
        
        pat = re.compile(r'(\[/?[a-z]+[^]]*\])', re.IGNORECASE)
        parts = pat.split(raw_text)
        
        sz = 10
        cl = 'black'
        is_bold = False
        is_italic = False
        is_underline = False
        
        for p in parts:
            if not p: continue
            lp = p.lower()
            
            if lp.startswith('[size='):
                m = re.search(r'=(\d+)', p)
                if m: sz = int(m.group(1))
            elif lp == '[/size]':
                sz = 10
            elif lp.startswith('[color='):
                m = re.search(r'=([^]]+)', p)
                if m: cl = m.group(1)
            elif lp == '[/color]':
                cl = 'black'
            elif lp == '[b]':
                is_bold = True
            elif lp == '[/b]':
                is_bold = False
            elif lp == '[i]':
                is_italic = True
            elif lp == '[/i]':
                is_italic = False
            elif lp == '[u]':
                is_underline = True
            elif lp == '[/u]':
                is_underline = False
            elif lp == '[br]':
                s = self.preview.index(tk.INSERT)
                self.preview.insert(tk.END, '\n')
            else:
                if p.strip() or '\n' in p:
                    s = self.preview.index(tk.INSERT)
                    self.preview.insert(tk.END, p)
                    e = self.preview.index(tk.INSERT)
                    
                    font_weight = 'bold' if is_bold else 'normal'
                    font_slant = 'italic' if is_italic else 'roman'
                    font_underline = True if is_underline else False
                    
                    tag_name = f'u_{sz}_{cl}_{font_weight}_{font_slant}_{font_underline}'
                    font = ('微软雅黑', sz, font_weight, font_slant)
                    self.preview.tag_configure(tag_name, font=font, foreground=cl, underline=font_underline)
                    self.preview.tag_add(tag_name, s, e)

        self.preview.config(state=tk.DISABLED)

    def _apply_size(self, event=None):
        sz = self.size_var.get()
        if sz == '选择': return
        try:
            s, e = self.editor.index(tk.SEL_FIRST), self.editor.index(tk.SEL_LAST)
            raw = self.editor.get(s, e)
            clean = strip_ubb_tags(raw)
            new = f'[size={sz}]{clean}[/size]'
            self.editor.delete(s, e)
            self.editor.insert(s, new)
            self._update_preview()
        except tk.TclError: messagebox.showwarning('提示', '请先选择文字')

    def _apply_custom_color(self):
        c = colorchooser.askcolor(title='选择颜色')
        if not c[1]: return
        try:
            s, e = self.editor.index(tk.SEL_FIRST), self.editor.index(tk.SEL_LAST)
            raw = self.editor.get(s, e)
            clean = strip_ubb_tags(raw)
            hex_color = c[1].upper()
            new = f'[color={hex_color}]{clean}[/color]'
            self.editor.delete(s, e)
            self.editor.insert(s, new)
            self._update_preview()
        except tk.TclError: messagebox.showwarning('提示', '请先选择文字')

    def _save_desc(self):
        if not self.current_row or not self.file_path: return messagebox.showwarning('警告', '请先选择并打开文件')
        try:
            content = self.editor.get('1.0', tk.END).strip()
            new_desc = content.replace('\r\n', '[br]').replace('\n', '[br]')
            row = self.current_row['row']
            
            if self.file_type == '.xls':
                rb = xlrd.open_workbook(self.file_path, formatting_info=False)
                wb = xl_copy(rb)
                wb.get_sheet(0).write(row, 25, new_desc)
                tmp = self.file_path + '.tmp'
                wb.save(tmp)
                os.replace(tmp, self.file_path)
                self.xlrd_book = xlrd.open_workbook(self.file_path, formatting_info=False)
            else:
                self.wb.active.cell(row=row, column=26).value = new_desc
                self.wb.save(self.file_path)
            
            self.current_row['desc'] = new_desc
            self.original_text.config(state=tk.NORMAL)
            self.original_text.delete('1.0', tk.END)
            self.original_text.insert('1.0', new_desc)
            self.original_text.config(state=tk.DISABLED)
            messagebox.showinfo('成功', '已保存回原文件')
        except Exception as e: messagebox.showerror('错误', str(e))

    def _show_log(self):
        w = tk.Toplevel(self.root)
        w.title('日志')
        w.geometry('800x500')
        t = tk.Text(w, wrap=tk.WORD)
        t.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        t.insert('1.0', self.logger.get_log_content())
        t.config(state=tk.DISABLED)
        f = tk.Frame(w)
        f.pack(fill=tk.X, padx=5, pady=5)
        tk.Button(f, text='刷新', command=lambda: (t.config(state=tk.NORMAL), t.delete('1.0',tk.END), t.insert('1.0',self.logger.get_log_content()), t.config(state=tk.DISABLED))).pack(side=tk.LEFT, padx=5)
        tk.Button(f, text='关闭', command=w.destroy).pack(side=tk.RIGHT, padx=5)

    def run(self):
        self.root.mainloop()

if __name__ == '__main__':
    ItemNoteEditor().run()
