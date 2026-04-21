# -*- coding: utf-8 -*-
"""
怪物编辑器 - Monster Editor
用于编辑 Monster.xls 文件中的怪物属性数据
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import xlrd
from xlutils.copy import copy as xl_copy
import os
import sys
import re
import logging


class MonsterEditor:

    def __init__(self):
        self.root = tk.Tk()
        self.root.title('怪物编辑器')
        self.root.geometry('1400x900')

        # 初始化日志
        self.logger = logging.getLogger('MonsterEditor')
        self.logger.setLevel(logging.DEBUG)
        log_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'monster_editor.log')
        fh = logging.FileHandler(log_file, encoding='utf-8')
        fh.setLevel(logging.DEBUG)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        fh.setFormatter(formatter)
        if not self.logger.handlers:
            self.logger.addHandler(fh)

        # 自动保存防抖定时器
        self.auto_save_timer = None

        # 基础路径设置
        self.base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        self.file_path = None
        self.file_type = None
        self.xlrd_book = None
        self.wb = None
        self.monsters = []
        self.current_row = None

        # 属性ID映射
        self.attr_mapping = {
            '生命值': 1,
            '攻击力': 23,
            '命中': 50,
            '防御': 52
        }

        self._setup_ui()
        self._try_load_default()

    def _setup_ui(self):
        """设置用户界面"""
        # 顶部工具栏
        toolbar = tk.Frame(self.root, bg='#f0f0f0', height=40)
        toolbar.pack(fill=tk.X, padx=5, pady=5)

        self.btn_open = tk.Button(toolbar, text='打开文件', command=self.open_file, 
                                  bg='#4CAF50', fg='white', font=('微软雅黑', 9))
        self.btn_open.pack(side=tk.LEFT, padx=5)

        self.btn_reload = tk.Button(toolbar, text='重新读取文件', command=self._reload_file, 
                                    bg='#FF9800', fg='white', font=('微软雅黑', 9))
        self.btn_reload.pack(side=tk.LEFT, padx=5)

        self.lbl_file = tk.Label(toolbar, text='未选择文件', bg='#f0f0f0',
                                 fg='#666', font=('微软雅黑', 9))
        self.lbl_file.pack(side=tk.LEFT, padx=10)

        # 保存按钮已移至编辑区域右下角，工具栏不再显示

        # 主容器
        main_frame = tk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 左侧：怪物列表
        left_frame = tk.Frame(main_frame, width=350)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 5))
        left_frame.pack_propagate(False)

        # 搜索框
        search_frame = tk.Frame(left_frame)
        search_frame.pack(fill=tk.X, pady=(0, 5))
        
        tk.Label(search_frame, text='搜索:', font=('微软雅黑', 10)).pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self._filter_monsters)
        tk.Entry(search_frame, textvariable=self.search_var, font=('微软雅黑', 9)).pack(
            fill=tk.X, expand=True, padx=(5, 0))

        # 列表容器
        list_frame = tk.Frame(left_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(list_frame, columns=('ID', 'Name'), show='headings', height=40)
        self.tree.heading('ID', text='ID')
        self.tree.heading('Name', text='怪物名称')
        self.tree.column('ID', width=60, minwidth=50)
        self.tree.column('Name', width=250, minwidth=200)

        tree_scroll = tk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.tree.bind('<<TreeviewSelect>>', self._on_monster_select)

        # 右侧：编辑区域
        right_frame = tk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # 欢迎界面
        self.welcome_frame = tk.Frame(right_frame)
        self.welcome_frame.pack(expand=True)

        tk.Label(self.welcome_frame, text='怪物编辑器',
                font=('微软雅黑', 16, 'bold'), fg='#333').pack(pady=20)

        tk.Label(self.welcome_frame, text='默认读取 ..\\..\\Mir200\\Envir\\data\\Monster.xls',
                font=('微软雅黑', 10), fg='#666').pack()

        tk.Label(self.welcome_frame, text='选择左侧怪物后可编辑等级、移速、经验值等属性',
                font=('微软雅黑', 9), fg='#FF9800').pack(pady=10)

        # 编辑容器（初始隐藏）
        self.edit_container = tk.Frame(right_frame)
        
        self._setup_edit_area()

    def _setup_edit_area(self):
        """设置编辑区域"""
        # 顶部信息
        top_frame = tk.Frame(self.edit_container)
        top_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Label(top_frame, text='当前编辑:', font=('微软雅黑', 11, 'bold'),
                fg='#2196F3').pack(anchor=tk.W)
        self.lbl_monster_name = tk.Label(top_frame, text='', font=('微软雅黑', 11), fg='#333')
        self.lbl_monster_name.pack(anchor=tk.W)

        # 使用Canvas实现可滚动区域
        canvas = tk.Canvas(self.edit_container)
        scrollbar = ttk.Scrollbar(self.edit_container, orient="vertical", command=canvas.yview)
        scrollable_wrapper = tk.Frame(canvas)

        scrollable_wrapper.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_wrapper, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # 内容容器（放在scrollable_wrapper中）
        content_frame = tk.Frame(scrollable_wrapper)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 基础属性字段
        self._setup_basic_fields(content_frame)

        # 属性编辑区
        self._setup_advanced_fields(content_frame)

        # 右下角保存按钮（缩小到0.7倍）
        save_frame = tk.Frame(scrollable_wrapper)
        save_frame.pack(fill=tk.X, padx=10, pady=10)

        self.btn_save_edit = tk.Button(save_frame, text='💾 保存修改', command=self._save_data,
                                       bg='#2196F3', fg='white', font=('微软雅黑', 8),
                                       state=tk.NORMAL, width=11, height=1)
        self.btn_save_edit.pack(side=tk.RIGHT)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def _setup_basic_fields(self, parent):
        """设置基础属性编辑区"""
        # 基础属性容器
        basic_container = tk.Frame(parent)
        basic_container.pack(fill=tk.X, padx=10, pady=5)

        # 标题
        tk.Label(basic_container, text='基础属性', font=('微软雅黑', 12, 'bold'), 
                fg='#2196F3').pack(anchor=tk.W, pady=5)

        # 显示名称字段（放在最上面）
        showname_frame = tk.Frame(basic_container)
        showname_frame.pack(fill=tk.X, pady=3)

        tk.Label(showname_frame, text='显示名称 (ShowName)', font=('微软雅黑', 10),
                width=20, anchor='w').pack(side=tk.LEFT)
        self.show_name_entry = tk.Entry(showname_frame, font=('微软雅黑', 10), width=30)
        self.show_name_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)

        # 基础字段
        fields = [
            ('等级 (Level)', 'level'),
            ('经验值 (Exp)', 'exp'),
            ('历练值 (HonorExp)', 'honor_exp'),
        ]

        self.basic_entries = {}
        for label_text, key in fields:
            field_frame = tk.Frame(basic_container)
            field_frame.pack(fill=tk.X, pady=3)

            tk.Label(field_frame, text=label_text, font=('微软雅黑', 10),
                    width=20, anchor='w').pack(side=tk.LEFT)

            entry = tk.Entry(field_frame, font=('微软雅黑', 10), width=30)
            entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
            entry.bind('<FocusOut>', lambda e: self._auto_save())
            self.basic_entries[key] = entry

        # 显示名称字段绑定事件
        self.show_name_entry.bind('<FocusOut>', lambda e: self._auto_save())

        # 移动速度拆分字段
        movespeed_frame = tk.Frame(basic_container)
        movespeed_frame.pack(fill=tk.X, pady=3)

        tk.Label(movespeed_frame, text='移动速度 (MoveSpeed)', font=('微软雅黑', 10),
                width=20, anchor='w').pack(side=tk.LEFT)

        speed_values_frame = tk.Frame(movespeed_frame)
        speed_values_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        tk.Label(speed_values_frame, text='步长(毫米/步):', font=('微软雅黑', 9)).pack(side=tk.LEFT, padx=(0, 2))
        self.movespeed_entry1 = tk.Entry(speed_values_frame, font=('微软雅黑', 10), width=10)
        self.movespeed_entry1.pack(side=tk.LEFT, padx=2)
        self.movespeed_entry1.bind('<FocusOut>', lambda e: self._auto_save())

        tk.Label(speed_values_frame, text='步频(毫秒/步):', font=('微软雅黑', 9)).pack(side=tk.LEFT, padx=(10, 2))
        self.movespeed_entry2 = tk.Entry(speed_values_frame, font=('微软雅黑', 10), width=10)
        self.movespeed_entry2.pack(side=tk.LEFT, padx=2)
        self.movespeed_entry2.bind('<FocusOut>', lambda e: self._auto_save())

        # 银两拆分字段
        silver_frame = tk.Frame(basic_container)
        silver_frame.pack(fill=tk.X, pady=3)

        tk.Label(silver_frame, text='银两 (Silver)', font=('微软雅黑', 10),
                width=20, anchor='w').pack(side=tk.LEFT)

        silver_values_frame = tk.Frame(silver_frame)
        silver_values_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)

        tk.Label(silver_values_frame, text='最小值:', font=('微软雅黑', 9)).pack(side=tk.LEFT, padx=(0, 2))
        self.silver_entry_min = tk.Entry(silver_values_frame, font=('微软雅黑', 10), width=10)
        self.silver_entry_min.pack(side=tk.LEFT, padx=2)
        self.silver_entry_min.bind('<FocusOut>', lambda e: self._auto_save())

        tk.Label(silver_values_frame, text='最大值:', font=('微软雅黑', 9)).pack(side=tk.LEFT, padx=(10, 2))
        self.silver_entry_max = tk.Entry(silver_values_frame, font=('微软雅黑', 10), width=10)
        self.silver_entry_max.pack(side=tk.LEFT, padx=2)
        self.silver_entry_max.bind('<FocusOut>', lambda e: self._auto_save())

        # 分隔线
        sep = tk.Frame(basic_container, height=2, bg='#ddd')
        sep.pack(fill=tk.X, pady=10)

        # 属性字符串显示区（只读）
        attr_frame = tk.Frame(basic_container)
        attr_frame.pack(fill=tk.X, pady=5)

        tk.Label(attr_frame, text='属性字符串 (只读):', font=('微软雅黑', 10, 'bold'), 
                fg='#FF5722').pack(anchor=tk.W)
        tk.Label(attr_frame, text='格式: 属性ID#属性值|属性ID#属性值', 
                font=('微软雅黑', 8), fg='#666').pack(anchor=tk.W)

        self.attr_string_var = tk.StringVar()
        
        attr_input_frame = tk.Frame(attr_frame)
        attr_input_frame.pack(fill=tk.X, pady=5)
        
        self.attr_string_entry = tk.Entry(attr_input_frame, textvariable=self.attr_string_var,
                                          font=('微软雅黑', 10), state='readonly')
        self.attr_string_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        tk.Button(attr_input_frame, text='解析', command=self._parse_attr_string, 
                 bg='#4CAF50', fg='white', font=('微软雅黑', 9), width=8).pack(side=tk.RIGHT)

    def _setup_advanced_fields(self, parent):
        """设置高级属性编辑区（属性ID专用分区）"""
        # 属性编辑容器
        attr_container = tk.Frame(parent)
        attr_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # 快速添加属性按钮
        btn_frame = tk.Frame(attr_container)
        btn_frame.pack(fill=tk.X, pady=5)

        tk.Label(btn_frame, text='快速添加:', font=('微软雅黑', 9, 'bold')).pack(side=tk.LEFT)
        
        preset_frame = tk.Frame(btn_frame)
        preset_frame.pack(side=tk.LEFT, padx=5)

        for attr_name, attr_id in self.attr_mapping.items():
            btn = tk.Button(preset_frame, text=f'{attr_name}({attr_id})', 
                           command=lambda aid=attr_id: self._add_attr_id(aid),
                           bg='#E3F2FD', fg='#1976D2', font=('微软雅黑', 9), width=15)
            btn.pack(side=tk.LEFT, padx=2)

        # 属性列表
        list_container = tk.Frame(attr_container)
        list_container.pack(fill=tk.BOTH, expand=True, pady=5)

        tk.Label(list_container, text='已添加的属性 (双击可编辑):', font=('微软雅黑', 10, 'bold')).pack(anchor=tk.W)

        # 属性列表Treeview
        self.attr_tree = ttk.Treeview(list_container, columns=('ID', 'Name', 'Value'), 
                                      show='headings', height=10)
        self.attr_tree.heading('ID', text='属性ID')
        self.attr_tree.heading('Name', text='属性名称')
        self.attr_tree.heading('Value', text='属性值')
        self.attr_tree.column('ID', width=100)
        self.attr_tree.column('Name', width=150)
        self.attr_tree.column('Value', width=200)

        attr_scroll = tk.Scrollbar(list_container, orient=tk.VERTICAL, command=self.attr_tree.yview)
        self.attr_tree.configure(yscrollcommand=attr_scroll.set)

        self.attr_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        attr_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # 绑定双击事件
        self.attr_tree.bind('<Double-Button-1>', self._edit_attr_value)

        # 操作按钮
        attr_btn_frame = tk.Frame(attr_container)
        attr_btn_frame.pack(fill=tk.X, pady=5)

        tk.Button(attr_btn_frame, text='添加自定义属性', command=self._add_custom_attr, 
                 bg='#4CAF50', fg='white', font=('微软雅黑', 9)).pack(side=tk.LEFT, padx=5)
        tk.Button(attr_btn_frame, text='删除选中', command=self._delete_selected_attr, 
                 bg='#F44336', fg='white', font=('微软雅黑', 9)).pack(side=tk.LEFT, padx=5)
        tk.Button(attr_btn_frame, text='清空所有', command=self._clear_all_attrs, 
                 bg='#FF9800', fg='white', font=('微软雅黑', 9)).pack(side=tk.LEFT, padx=5)

    def _try_load_default(self):
        """尝试加载默认文件"""
        default_xls = os.path.join(self.base_dir, '..', '..', 'Mirserver', 'Mir200', 'Envir', 'data', 'Monster.xls')
        default_xlsx = os.path.join(self.base_dir, '..', '..', 'Mirserver', 'Mir200', 'Envir', 'data', 'Monster.xlsx')

        if os.path.exists(default_xlsx):
            self._load_file(default_xlsx)
        elif os.path.exists(default_xls):
            self._load_file(default_xls)

    def open_file(self):
        """打开文件对话框"""
        file_path = filedialog.askopenfilename(
            title='选择Monster文件',
            filetypes=[('Excel文件', '*.xls *.xlsx')],
            initialdir=self.base_dir
        )
        if file_path:
            self._load_file(file_path)

    def _reload_file(self):
        """重新加载当前文件"""
        if not self.file_path:
            return messagebox.showwarning('警告', '请先打开文件')
        try:
            self._load_file(self.file_path)
            messagebox.showinfo('成功', '文件已重新读取')
        except Exception as e:
            messagebox.showerror('错误', str(e))

    def _load_file(self, file_path):
        """加载Excel文件"""
        try:
            self.logger.info(f'开始加载文件: {file_path}')
            self.file_path = file_path
            self.file_type = os.path.splitext(file_path)[1].lower()

            if self.file_type == '.xls':
                self._load_xls(file_path)
            elif self.file_type == '.xlsx':
                self._load_xlsx(file_path)
            else:
                return messagebox.showwarning('警告', '仅支持 .xls 或 .xlsx 格式')

            self.lbl_file.config(text=f'{os.path.basename(file_path)} ({self.file_type}) - {len(self.monsters)}个怪物')
            self.current_row = None
            self.logger.info(f'加载完成，共 {len(self.monsters)} 个怪物')

            # 自动选择第一个怪物
            if self.monsters:
                self.tree.selection_set(self.tree.get_children()[0])
                self.tree.focus(self.tree.get_children()[0])
                self.tree.see(self.tree.get_children()[0])
        except Exception as e:
            self.logger.error(f'加载文件失败: {str(e)}', exc_info=True)
            messagebox.showerror('错误', f'加载文件失败: {str(e)}')

    def _load_xls(self, fp):
        """加载.xls文件"""
        self.xlrd_book = xlrd.open_workbook(fp, formatting_info=False)
        sh = self.xlrd_book.sheet_by_index(0)

        # 查找表头行
        hdr = self._find_header(sh)

        # 解析数据 (B列=索引1为ShowName, C列=索引2为Name)
        self.monsters = []
        for r in range(hdr + 1, sh.nrows):
            mid = self._try_id(sh.cell_value(r, 0))
            if mid is None:
                continue
            name = sh.cell_value(r, 2)  # C列
            if not name:
                continue
            show_name = sh.cell_value(r, 1)  # B列
            attr_str = sh.cell_value(r, 5) if sh.ncols > 5 else ''  # F列 Props
            self.monsters.append({
                'row': r,
                'id': mid,
                'name': str(name),
                'show_name': str(show_name) if show_name else '',
                'attr_str': str(attr_str) if attr_str else ''
            })

        self._populate_tree()

    def _load_xlsx(self, fp):
        """加载.xlsx文件"""
        self.wb = openpyxl.load_workbook(fp)
        ws = self.wb.active

        hdr = self._find_header_ws(ws)

        self.monsters = []
        for r in range(hdr + 1, ws.max_row + 1):
            mid = self._try_id(ws.cell(row=r, column=1).value)
            if mid is None:
                continue
            name = ws.cell(row=r, column=3).value  # C列
            if not name:
                continue
            show_name = ws.cell(row=r, column=2).value or ''  # B列
            attr_str = ws.cell(row=r, column=6).value or ''  # F列 Props
            self.monsters.append({
                'row': r,
                'id': mid,
                'name': str(name),
                'show_name': str(show_name),
                'attr_str': str(attr_str) if attr_str else ''
            })

        self._populate_tree()

    def _find_header(self, sh):
        """查找.xls文件的表头行"""
        for r in range(min(10, sh.nrows)):
            if any('id' in str(sh.cell_value(r, c)).lower() for c in range(min(3, sh.ncols))):
                return r
        return 0

    def _find_header_ws(self, ws):
        """查找.xlsx文件的表头行"""
        for r in range(1, min(11, ws.max_row + 1)):
            if any('id' in str(ws.cell(row=r, column=c).value or '').lower() for c in range(1, 4)):
                return r
        return 1

    def _try_id(self, v):
        """尝试将值转换为ID"""
        try:
            return int(float(v))
        except:
            return None

    def _populate_tree(self):
        """填充怪物列表"""
        for i in self.tree.get_children():
            self.tree.delete(i)
        for m in self.monsters:
            self.tree.insert('', tk.END, values=(m['id'], m['name']))

    def _filter_monsters(self, *args):
        """过滤怪物列表"""
        st = self.search_var.get().lower()
        for i in self.tree.get_children():
            self.tree.delete(i)
        for m in self.monsters:
            if st in str(m['name']).lower() or st in str(m['id']).lower() or st in str(m.get('show_name', '')).lower():
                self.tree.insert('', tk.END, values=(m['id'], m['name']))

    def _on_monster_select(self, event):
        """怪物选择事件"""
        sel = self.tree.selection()
        if not sel:
            return
        
        mid = int(self.tree.item(sel[0])['values'][0])
        for m in self.monsters:
            if m['id'] == mid:
                self.current_row = m
                self._show_editor(m)
                return

    def _show_editor(self, monster):
        """显示编辑器"""
        # 隐藏欢迎界面，显示编辑界面
        self.welcome_frame.pack_forget()
        self.edit_container.pack(fill=tk.BOTH, expand=True)

        # 更新怪物名称显示
        self.lbl_monster_name.config(text=f'{monster["name"]} (ID: {monster["id"]})')

        # 先清空属性列表，再加载数据（确保执行顺序正确）
        for i in self.attr_tree.get_children():
            self.attr_tree.delete(i)

        # 读取Excel中的数据（包含自动解析属性字符串）
        row = monster['row']
        self._load_monster_data(row)

    def _load_monster_data(self, row):
        """加载怪物数据到编辑区"""
        try:
            if self.file_type == '.xls':
                sh = self.xlrd_book.sheet_by_index(0)
                # D列=等级(索引3), F列=Props(索引5), G列=移速(索引6), I列=经验(索引8), J列=历练(索引9), K列=银两(索引10)
                level = sh.cell_value(row, 3) if sh.ncols > 3 else ''
                movespeed = sh.cell_value(row, 6) if sh.ncols > 6 else ''
                exp = sh.cell_value(row, 8) if sh.ncols > 8 else ''
                honor_exp = sh.cell_value(row, 9) if sh.ncols > 9 else ''
                silver = sh.cell_value(row, 10) if sh.ncols > 10 else ''
                attr_str = sh.cell_value(row, 5) if sh.ncols > 5 else ''  # F列 Props
            else:
                ws = self.wb.active
                # D列=等级(列4), F列=Props(列6), G列=移速(列7), I列=经验(列9), J列=历练(列10), K列=银两(列11)
                level = ws.cell(row=row, column=4).value or ''
                movespeed = ws.cell(row=row, column=7).value or ''
                exp = ws.cell(row=row, column=9).value or ''
                honor_exp = ws.cell(row=row, column=10).value or ''
                silver = ws.cell(row=row, column=11).value or ''
                attr_str = ws.cell(row=row, column=6).value or ''  # F列 Props

            # 加载显示名称（从缓存的数据中获取，避免重新读取文件导致数据丢失）
            show_name = self.current_row.get('show_name', '') if self.current_row else ''

            self.show_name_entry.delete(0, tk.END)
            self.show_name_entry.insert(0, str(show_name) if show_name else '')
            self.logger.info(f'加载显示名称: {show_name}')

            # 填充基础字段
            self.basic_entries['level'].delete(0, tk.END)
            self.basic_entries['level'].insert(0, level)

            # 解析移动速度格式 {值1,值2}
            self.movespeed_entry1.delete(0, tk.END)
            self.movespeed_entry2.delete(0, tk.END)
            movespeed_str = str(movespeed).strip()
            if movespeed_str.startswith('{') and movespeed_str.endswith('}'):
                # 移除大括号，按逗号分割
                content = movespeed_str[1:-1]
                parts = content.split(',')
                if len(parts) >= 2:
                    self.movespeed_entry1.insert(0, parts[0].strip())
                    self.movespeed_entry2.insert(0, parts[1].strip())
                elif len(parts) == 1:
                    self.movespeed_entry1.insert(0, parts[0].strip())
            else:
                # 如果不是{}格式，直接填充
                self.movespeed_entry1.insert(0, movespeed_str)

            self.basic_entries['exp'].delete(0, tk.END)
            self.basic_entries['exp'].insert(0, exp)

            self.basic_entries['honor_exp'].delete(0, tk.END)
            self.basic_entries['honor_exp'].insert(0, honor_exp)

            # 解析银两格式 最小值#最大值
            self.silver_entry_min.delete(0, tk.END)
            self.silver_entry_max.delete(0, tk.END)
            silver_str = str(silver).strip()
            if '#' in silver_str:
                parts = silver_str.split('#')
                if len(parts) >= 2:
                    self.silver_entry_min.insert(0, parts[0].strip())
                    self.silver_entry_max.insert(0, parts[1].strip())
                elif len(parts) == 1:
                    self.silver_entry_min.insert(0, parts[0].strip())
            else:
                # 如果不是#格式，直接填充到最小值
                self.silver_entry_min.insert(0, silver_str)

            # 属性字符串
            self.attr_string_var.set(str(attr_str) if attr_str else '')

            # 缓存原始属性字符串，防止丢失
            if self.current_row:
                self.current_row['attr_str'] = str(attr_str) if attr_str else ''

            # 延迟解析属性字符串，确保UI已完全加载
            attr_str_val = str(attr_str) if attr_str else ''
            self.root.after(50, lambda val=attr_str_val: self._parse_attr_string_to_tree(val))

        except Exception as e:
            messagebox.showerror('错误', f'加载数据失败: {str(e)}')

    def _parse_attr_string_to_tree(self, attr_str):
        """解析属性字符串并填充到属性列表"""
        # 清空属性列表
        for i in self.attr_tree.get_children():
            self.attr_tree.delete(i)

        if not attr_str:
            return

        # 格式: 属性ID#属性值|属性ID#属性值
        parts = attr_str.split('|')
        for part in parts:
            part = part.strip()
            if '#' in part:
                try:
                    attr_id, attr_value = part.split('#', 1)
                    attr_id = int(attr_id)
                    attr_value = attr_value.strip()
                    
                    # 查找属性名称
                    attr_name = self._get_attr_name_by_id(attr_id)
                    
                    self.attr_tree.insert('', tk.END, values=(attr_id, attr_name, attr_value))
                except:
                    pass

    def _get_attr_name_by_id(self, attr_id):
        """根据属性ID获取名称"""
        for name, aid in self.attr_mapping.items():
            if aid == attr_id:
                return name
        return f'未知属性({attr_id})'

    def _get_attr_id_by_name(self, attr_name):
        """根据属性名称获取ID"""
        return self.attr_mapping.get(attr_name, None)

    def _parse_attr_string(self):
        """解析属性字符串到属性列表"""
        attr_str = self.attr_string_var.get()
        self._parse_attr_string_to_tree(attr_str)

    def _generate_attr_string(self):
        """从属性列表生成属性字符串"""
        parts = []
        for item in self.attr_tree.get_children():
            values = self.attr_tree.item(item)['values']
            attr_id = values[0]
            attr_value = values[2]
            parts.append(f'{attr_id}#{attr_value}')
        
        attr_string = '|'.join(parts)
        self.attr_string_var.set(attr_string)

    def _edit_attr_value(self, event):
        """双击编辑属性值"""
        selection = self.attr_tree.selection()
        if not selection:
            return
        
        item = selection[0]
        values = self.attr_tree.item(item)['values']
        attr_id = values[0]
        attr_name = values[1]
        current_value = values[2]
        
        # 创建编辑对话框
        dialog = tk.Toplevel(self.root)
        dialog.title(f'编辑属性值 - {attr_name}')
        dialog.geometry('300x150')
        dialog.transient(self.root)
        dialog.grab_set()

        tk.Label(dialog, text=f'{attr_name} (ID: {attr_id}):', font=('微软雅黑', 10)).pack(pady=10)
        
        value_entry = tk.Entry(dialog, font=('微软雅黑', 12), width=20)
        value_entry.insert(0, str(current_value))
        value_entry.pack(pady=5)
        value_entry.focus_set()
        value_entry.select_range(0, tk.END)

        def on_ok():
            new_value = value_entry.get()
            self.attr_tree.item(item, values=(attr_id, attr_name, new_value))
            self._generate_attr_string()  # 自动更新属性字符串
            self._auto_save()  # 自动保存
            dialog.destroy()

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text='确定', command=on_ok, bg='#4CAF50', 
                 fg='white', font=('微软雅黑', 9)).pack(side=tk.LEFT, padx=10)
        tk.Button(btn_frame, text='取消', command=dialog.destroy, bg='#9E9E9E', 
                 fg='white', font=('微软雅黑', 9)).pack(side=tk.LEFT, padx=10)

        # 居中显示
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f'+{x}+{y}')
        
        # 绑定回车键
        value_entry.bind('<Return>', lambda e: on_ok())

    def _add_attr_id(self, attr_id):
        """添加预设属性"""
        attr_name = self._get_attr_name_by_id(attr_id)
        # 检查是否已存在
        for item in self.attr_tree.get_children():
            if self.attr_tree.item(item)['values'][0] == attr_id:
                messagebox.showinfo('提示', f'{attr_name}已存在')
                return
        
        self.attr_tree.insert('', tk.END, values=(attr_id, attr_name, 0))
        self._generate_attr_string()  # 自动更新属性字符串
        self._auto_save()  # 自动保存

    def _add_custom_attr(self):
        """添加自定义属性"""
        dialog = tk.Toplevel(self.root)
        dialog.title('添加属性')
        dialog.geometry('350x200')
        dialog.transient(self.root)
        dialog.grab_set()

        tk.Label(dialog, text='属性ID:', font=('微软雅黑', 10)).pack(pady=5)
        id_entry = tk.Entry(dialog, font=('微软雅黑', 10))
        id_entry.pack(pady=5)

        tk.Label(dialog, text='属性值:', font=('微软雅黑', 10)).pack(pady=5)
        value_entry = tk.Entry(dialog, font=('微软雅黑', 10))
        value_entry.pack(pady=5)

        def on_ok():
            try:
                attr_id = int(id_entry.get())
                attr_value = value_entry.get()
                attr_name = self._get_attr_name_by_id(attr_id)
                self.attr_tree.insert('', tk.END, values=(attr_id, attr_name, attr_value))
                self._generate_attr_string()  # 自动更新属性字符串
                self._auto_save()  # 自动保存
                dialog.destroy()
            except:
                messagebox.showerror('错误', '请输入有效的属性ID和值')

        btn_frame = tk.Frame(dialog)
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text='确定', command=on_ok, bg='#4CAF50',
                 fg='white', font=('微软雅黑', 9)).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text='取消', command=dialog.destroy, bg='#9E9E9E',
                 fg='white', font=('微软雅黑', 9)).pack(side=tk.LEFT, padx=5)

        # 居中显示
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (dialog.winfo_width() // 2)
        y = (dialog.winfo_screenheight() // 2) - (dialog.winfo_height() // 2)
        dialog.geometry(f'+{x}+{y}')

    def _delete_selected_attr(self):
        """删除选中的属性"""
        selected = self.attr_tree.selection()
        if not selected:
            messagebox.showinfo('提示', '请先选择要删除的属性')
            return

        for item in selected:
            self.attr_tree.delete(item)
        self._generate_attr_string()  # 自动更新属性字符串
        self._auto_save()  # 自动保存

    def _clear_all_attrs(self):
        """清空所有属性"""
        if messagebox.askyesno('确认', '确定要清空所有属性吗？'):
            for item in self.attr_tree.get_children():
                self.attr_tree.delete(item)
            self._generate_attr_string()  # 自动更新属性字符串
            self._auto_save()  # 自动保存

    def _auto_save(self):
        """自动保存（带防抖的静默保存）"""
        # 取消之前的定时器
        if self.auto_save_timer:
            self.root.after_cancel(self.auto_save_timer)
        
        # 延迟200ms执行保存，避免频繁保存
        self.auto_save_timer = self.root.after(200, self._do_auto_save)
    
    def _do_auto_save(self):
        """执行实际的自动保存"""
        if not self.current_row or not self.file_path:
            self.logger.warning('自动保存跳过: 未选择怪物或未打开文件')
            return

        try:
            row = self.current_row['row']
            self.logger.info(f'开始自动保存: {self.current_row["name"]} (行: {row})')

            # 生成属性字符串（如果属性列表为空，保留原始属性字符串）
            if self.attr_tree.get_children():
                self._generate_attr_string()
                attr_string = self.attr_string_var.get()
            else:
                attr_string = self.current_row.get('attr_str', '')

            # 合并移动速度为 {值1,值2} 格式，并格式化数字
            speed1 = self._format_number(self.movespeed_entry1.get().strip())
            speed2 = self._format_number(self.movespeed_entry2.get().strip())
            movespeed_value = f'{{{speed1},{speed2}}}'

            # 合并银两为 最小值#最大值 格式，并格式化数字
            silver_min = self._format_number(self.silver_entry_min.get().strip())
            silver_max = self._format_number(self.silver_entry_max.get().strip())
            silver_value = f'{silver_min}#{silver_max}'

            # 获取显示名称
            show_name = self.show_name_entry.get()

            # 格式化基础属性值
            level_value = self._format_number(self.basic_entries['level'].get())
            exp_value = self._format_number(self.basic_entries['exp'].get())
            honor_exp_value = self._format_number(self.basic_entries['honor_exp'].get())

            if self.file_type == '.xls':
                rb = xlrd.open_workbook(self.file_path, formatting_info=False)
                wb = xl_copy(rb)
                sheet = wb.get_sheet(0)

                sheet.write(row, 1, show_name)
                sheet.write(row, 3, level_value)
                sheet.write(row, 5, attr_string)
                sheet.write(row, 6, movespeed_value)
                sheet.write(row, 8, exp_value)
                sheet.write(row, 9, honor_exp_value)
                sheet.write(row, 10, silver_value)

                tmp = self.file_path + '.tmp'
                wb.save(tmp)
                os.replace(tmp, self.file_path)
                self.xlrd_book = xlrd.open_workbook(self.file_path, formatting_info=False)
            else:
                ws = self.wb.active
                ws.cell(row=row, column=2).value = show_name
                ws.cell(row=row, column=4).value = level_value
                ws.cell(row=row, column=6).value = attr_string
                ws.cell(row=row, column=7).value = movespeed_value
                ws.cell(row=row, column=9).value = exp_value
                ws.cell(row=row, column=10).value = honor_exp_value
                ws.cell(row=row, column=11).value = silver_value
                self.wb.save(self.file_path)

            # 更新数据模型
            self.current_row['show_name'] = show_name
            self.current_row['attr_str'] = attr_string
            self.logger.info(f'自动保存成功: {self.current_row["name"]}')
        except Exception as e:
            self.logger.error(f'自动保存失败: {str(e)}', exc_info=True)
            self.logger.error(f'错误详情 - 当前怪物: {self.current_row.get("name") if self.current_row else "None"}, 文件: {self.file_path}')

    def _format_number(self, value):
        """格式化数字：如果数字有.0则转换为整数，否则保持原样"""
        if value is None or value == '':
            return value
        try:
            # 尝试转换为浮点数
            num = float(value)
            # 如果是整数（小数部分为0），则返回整数形式
            if num == int(num):
                return str(int(num))
            else:
                # 否则保持原样（小数）
                return str(num)
        except (ValueError, TypeError):
            # 如果不是数字，保持原样
            return value

    def _save_data(self):
        """保存数据到Excel"""
        if not self.current_row or not self.file_path:
            return messagebox.showwarning('警告', '请先选择怪物并打开文件')

        try:
            row = self.current_row['row']
            self.logger.info(f'开始保存怪物数据: {self.current_row["name"]} (行: {row})')

            # 生成属性字符串
            self._generate_attr_string()
            attr_string = self.attr_string_var.get()
            self.logger.info(f'属性字符串: {attr_string}')

            # 合并移动速度为 {值1,值2} 格式，并格式化数字
            speed1 = self._format_number(self.movespeed_entry1.get().strip())
            speed2 = self._format_number(self.movespeed_entry2.get().strip())
            movespeed_value = f'{{{speed1},{speed2}}}'
            self.logger.info(f'移动速度: {movespeed_value}')

            # 合并银两为 最小值#最大值 格式，并格式化数字
            silver_min = self._format_number(self.silver_entry_min.get().strip())
            silver_max = self._format_number(self.silver_entry_max.get().strip())
            silver_value = f'{silver_min}#{silver_max}'
            self.logger.info(f'银两: {silver_value}')

            # 获取显示名称
            show_name = self.show_name_entry.get()
            self.logger.info(f'显示名称: {show_name}')

            # 格式化基础属性值
            level_value = self._format_number(self.basic_entries['level'].get())
            exp_value = self._format_number(self.basic_entries['exp'].get())
            honor_exp_value = self._format_number(self.basic_entries['honor_exp'].get())

            # 生成属性字符串（如果属性列表为空，保留原始属性字符串）
            if self.attr_tree.get_children():
                self._generate_attr_string()
                attr_string = self.attr_string_var.get()
            else:
                # 属性列表为空，使用原始值，避免丢失
                attr_string = self.current_row.get('attr_str', '')
            self.logger.info(f'属性字符串: {attr_string}')

            if self.file_type == '.xls':
                # .xls 文件使用 xlutils 保存（只修改值，不影响格式）
                rb = xlrd.open_workbook(self.file_path, formatting_info=False)
                wb = xl_copy(rb)
                sheet = wb.get_sheet(0)

                # B列=显示名称(索引1), D列=等级(索引3), F列=Props(索引5), G列=移速(索引6), I列=经验(索引8), J列=历练(索引9), K列=银两(索引10)
                sheet.write(row, 1, show_name)
                sheet.write(row, 3, level_value)
                sheet.write(row, 5, attr_string)  # F列 Props
                sheet.write(row, 6, movespeed_value)
                sheet.write(row, 8, exp_value)
                sheet.write(row, 9, honor_exp_value)
                sheet.write(row, 10, silver_value)

                tmp = self.file_path + '.tmp'
                wb.save(tmp)
                os.replace(tmp, self.file_path)
                self.xlrd_book = xlrd.open_workbook(self.file_path, formatting_info=False)
            else:
                # .xlsx 文件使用 openpyxl 保存（只修改值，不影响格式）
                ws = self.wb.active
                # B列=显示名称(列2), D列=等级(列4), F列=Props(列6), G列=移速(列7), I列=经验(列9), J列=历练(列10), K列=银两(列11)
                ws.cell(row=row, column=2).value = show_name
                ws.cell(row=row, column=4).value = level_value
                ws.cell(row=row, column=6).value = attr_string  # F列 Props
                ws.cell(row=row, column=7).value = movespeed_value
                ws.cell(row=row, column=9).value = exp_value
                ws.cell(row=row, column=10).value = honor_exp_value
                ws.cell(row=row, column=11).value = silver_value
                self.wb.save(self.file_path)

            # 更新数据模型
            self.current_row['show_name'] = show_name
            self.current_row['attr_str'] = attr_string
            self.logger.info('保存成功')

            messagebox.showinfo('成功', f'已保存: {self.current_row["name"]}')
        except Exception as e:
            self.logger.error(f'保存失败: {str(e)}', exc_info=True)
            messagebox.showerror('错误', f'保存失败: {str(e)}')

    def run(self):
        """运行应用"""
        self.root.mainloop()


if __name__ == '__main__':
    MonsterEditor().run()
