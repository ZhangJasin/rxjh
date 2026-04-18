# -*- coding: utf-8 -*-
"""
Item.xls 备注修改工具 (HTML + UBB 版)
支持两种富文本格式：
UBB: [size=xx] [color=xx] [br] [b] [i] [u]
HTML: <font size=xx color=xx> <b> <i> <u> <br>
所有标签使用方括号 [...] 或尖括号 <...> 格式
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser
from openpyxl import load_workbook
import xlrd
from xlutils.copy import copy as xl_copy
import os
import sys
import logging
import re
import html

# UBB 标签正则表达式 (匹配所有 [tag], [/tag], [tag=val])
UBB_TAG_RE = re.compile(r'\[/?[a-zA-Z]+[^]]*\]', re.IGNORECASE)
# HTML 标签正则表达式
HTML_TAG_RE = re.compile(r'</?[a-zA-Z][^>]*>', re.IGNORECASE)

def strip_ubb_tags(text):
    """移除文本中所有的 UBB 标签"""
    return UBB_TAG_RE.sub('', text)

def strip_html_tags(text):
    """移除文本中所有的 HTML 标签"""
    return HTML_TAG_RE.sub('', text)

def detect_format(text):
    """检测文本是 UBB 格式还是 HTML 格式"""
    has_ubb = bool(UBB_TAG_RE.search(text))
    has_html = bool(HTML_TAG_RE.search(text))
    if has_html:
        return 'html'
    elif has_ubb:
        return 'ubb'
    else:
        return 'plain'

class Logger:
    def __init__(self, log_file=None):
        self.log_file = log_file or os.path.join(os.path.dirname(os.path.abspath(__file__)), 'item_editor.log')
        self.logger = logging.getLogger('ItemEditor')
        self.logger.setLevel(logging.DEBUG)
        fh = logging.FileHandler(self.log_file, encoding='utf-8')
        fh.setLevel(logging.DEBUG)
        ch = logging.StreamHandler()
        ch.setLevel(logging.INFO)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        fh.setFormatter(formatter)
        ch.setFormatter(formatter)
        self.logger.addHandler(fh)
        self.logger.addHandler(ch)
    
    def info(self, msg): self.logger.info(msg)
    def warning(self, msg): self.logger.warning(msg)
    def error(self, msg): self.logger.error(msg)
    def get_log_content(self):
        try:
            with open(self.log_file, 'r', encoding='utf-8') as f: return f.read()
        except Exception: return '暂无日志'

class ItemNoteEditor:
    
    def __init__(self):
        self.root = tk.Tk()
        self.root.title('富文本编辑器')
        self.root.geometry('1200x850')
        
        self.base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        self.file_path = None
        self.file_type = None
        self.xlrd_book = None
        self.wb = None
        self.items = []
        self.current_row = None
        self.preview_timer = None
        
        self.logger = Logger()
        self.logger.info('=' * 50)
        self.logger.info('Item备注修改工具 (UBB版) 启动')
        self._setup_ui()
        self._try_load_default()

    def _setup_ui(self):
        toolbar = tk.Frame(self.root, bg='#f0f0f0', height=40)
        toolbar.pack(fill=tk.X, padx=5, pady=5)
        
        self.btn_open = tk.Button(toolbar, text='打开文件', command=self.open_file, bg='#4CAF50', fg='white', font=('微软雅黑', 9))
        self.btn_open.pack(side=tk.LEFT, padx=5)
        
        self.btn_reload = tk.Button(toolbar, text='重新读取文件', command=self._reload_file, bg='#FF9800', fg='white', font=('微软雅黑', 9))
        self.btn_reload.pack(side=tk.LEFT, padx=5)
        
        self.btn_refresh = tk.Button(toolbar, text='读取备注', command=self._refresh_desc, bg='#2196F3', fg='white', font=('微软雅黑', 9))
        self.btn_refresh.pack(side=tk.LEFT, padx=5)
        
        self.lbl_file = tk.Label(toolbar, text='未选择文件', bg='#f0f0f0', fg='#666', font=('微软雅黑', 9))
        self.lbl_file.pack(side=tk.LEFT, padx=10)
        
        self.btn_save = tk.Button(toolbar, text='保存备注', command=self._save_desc, bg='#4CAF50', fg='white', font=('微软雅黑', 9), state=tk.DISABLED)
        self.btn_save.pack(side=tk.RIGHT, padx=5)
        
        self.btn_log = tk.Button(toolbar, text='查看日志', command=self._show_log, bg='#607D8B', fg='white', font=('微软雅黑', 9))
        self.btn_log.pack(side=tk.RIGHT, padx=5)
        
        main_frame = tk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # 左侧：物品列表
        left_frame = tk.Frame(main_frame, width=300)
        left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 5))
        left_frame.pack_propagate(False)
        
        search_frame = tk.Frame(left_frame)
        search_frame.pack(fill=tk.X, pady=(0, 5))
        tk.Label(search_frame, text='搜索:', font=('微软雅黑', 10)).pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self._filter_items)
        tk.Entry(search_frame, textvariable=self.search_var, font=('微软雅黑', 9)).pack(fill=tk.X, expand=True, padx=(5, 0))
        
        list_frame = tk.Frame(left_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        self.tree = ttk.Treeview(list_frame, columns=('ID', 'Name'), show='headings', height=30)
        self.tree.heading('ID', text='ID')
        self.tree.heading('Name', text='道具名称')
        self.tree.column('ID', width=60, minwidth=50)
        self.tree.column('Name', width=200, minwidth=150)
        
        tree_scroll = tk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.tree.bind('<<TreeviewSelect>>', self._on_item_select)
        
        # 右侧：编辑区域
        right_frame = tk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        self.right_container = tk.Frame(right_frame)
        self.right_container.pack(fill=tk.BOTH, expand=True)
        
        welcome_frame = tk.Frame(self.right_container)
        welcome_frame.pack(expand=True)
        
        tk.Label(welcome_frame, text='富文本编辑器',
                font=('微软雅黑', 16, 'bold'), fg='#333').pack(pady=20)

        # 更新显示的路径
        tk.Label(welcome_frame, text='默认读取 ..\\..\\Mir200\\Envir\\data\\Item.xls',
                font=('微软雅黑', 10), fg='#666').pack()

        tk.Label(welcome_frame, text='修改备注后直接保存回原文件',
                font=('微软雅黑', 9), fg='#FF9800').pack(pady=10)
        tk.Label(welcome_frame, text='支持 UBB: [size=xx] [color=xx] [br] [b] [i] [u]',
                font=('微软雅黑', 9), fg='#4CAF50').pack(pady=5)
        tk.Label(welcome_frame, text='支持 HTML: <font size=xx color=xx> <b> <i> <u> <br>',
                font=('微软雅黑', 9), fg='#2196F3').pack(pady=5)

    def _try_load_default(self):
        # 修改后的默认路径：相对于 exe 目录的 ../../Mir200/Envir/data/
        default_xls = os.path.join(self.base_dir, '..', '..', 'Mir200', 'Envir', 'data', 'Item.xls')
        default_xlsx = os.path.join(self.base_dir, '..', '..', 'Mir200', 'Envir', 'data', 'Item.xlsx')

        self.logger.info(f'查找默认文件: {default_xls}')
        if os.path.exists(default_xlsx): 
            self._load_file(default_xlsx)
        elif os.path.exists(default_xls): 
            self._load_file(default_xls)
        else: 
            # 未找到默认文件时不提示,直接显示欢迎界面
            self.logger.info('未找到默认文件,等待用户手动选择')

    def open_file(self):
        file_path = filedialog.askopenfilename(
            title='选择Item文件',
            filetypes=[('Excel文件', '*.xls *.xlsx')],
            initialdir=self.base_dir
        )
        if file_path: self._load_file(file_path)

    def _reload_file(self):
        if not self.file_path: return messagebox.showwarning('警告', '请先打开文件')
        try:
            self._load_file(self.file_path)
            messagebox.showinfo('成功', '文件已重新读取')
        except Exception as e: messagebox.showerror('错误', str(e))

    def _refresh_desc(self):
        if not self.current_row: return messagebox.showwarning('警告', '请先选择道具')
        try:
            row = self.current_row['row']
            desc = self.xlrd_book.sheet_by_index(0).cell_value(row, 25) if self.file_type == '.xls' else self.wb.active.cell(row=row, column=26).value or ''
            self.current_row['desc'] = str(desc) if desc else ''

            display = self.current_row['desc'].replace('[br]', '\n').replace('<br>', '\n').replace('<br/>', '\n')
            self.editor.delete('1.0', tk.END)
            self.editor.insert('1.0', display)

            self._update_preview()
            messagebox.showinfo('成功', f'已重新读取：{self.current_row["name"]}')
        except Exception as e: messagebox.showerror('错误', str(e))

    def _load_file(self, file_path):
        try:
            self.file_path = file_path
            self.file_type = os.path.splitext(file_path)[1].lower()
            self.logger.info(f'加载: {file_path} ({self.file_type})')

            if self.file_type == '.xls': self._load_xls(file_path)
            elif self.file_type == '.xlsx': self._load_xlsx(file_path)
            else: return messagebox.showwarning('警告', '仅支持 .xls 或 .xlsx')

            self.lbl_file.config(text=f'{os.path.basename(file_path)} ({self.file_type}) - {len(self.items)}个物品')
            self.btn_save.config(state=tk.NORMAL)
            self.btn_refresh.config(state=tk.NORMAL)
            self.current_row = None
            
            # 自动选择第一个物品,直接显示编辑界面
            if self.items:
                self.tree.selection_set(self.tree.get_children()[0])
                self.tree.focus(self.tree.get_children()[0])
                self.tree.see(self.tree.get_children()[0])
        except Exception as e: messagebox.showerror('错误', str(e))

    def _load_xls(self, fp):
        self.xlrd_book = xlrd.open_workbook(fp, formatting_info=False)
        sh = self.xlrd_book.sheet_by_index(0)
        hdr = self._find_header(sh)
        self.items = []
        for r in range(hdr+1, sh.nrows):
            iid = self._try_id(sh.cell_value(r, 0))
            if iid is None: continue
            name = sh.cell_value(r, 1)
            if not name: continue
            self.items.append({'row': r, 'id': iid, 'name': str(name), 'desc': str(sh.cell_value(r, 25) or '')})
        self._populate_tree()

    def _load_xlsx(self, fp):
        self.wb = load_workbook(fp)
        ws = self.wb.active
        hdr = self._find_header_ws(ws)
        self.items = []
        for r in range(hdr+1, ws.max_row+1):
            iid = self._try_id(ws.cell(row=r, column=1).value)
            if iid is None: continue
            name = ws.cell(row=r, column=2).value
            if not name: continue
            self.items.append({'row': r, 'id': iid, 'name': str(name), 'desc': str(ws.cell(row=r, column=26).value or '')})
        self._populate_tree()

    def _find_header(self, sh):
        for r in range(min(10, sh.nrows)):
            if any('id' in str(sh.cell_value(r,c)).lower() for c in range(min(3, sh.ncols))): return r
        return 0

    def _find_header_ws(self, ws):
        for r in range(1, min(11, ws.max_row+1)):
            if any('id' in str(ws.cell(row=r, column=c).value or '').lower() for c in range(1,4)): return r
        return 1

    def _try_id(self, v):
        try: return int(float(v))
        except: return None

    def _populate_tree(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        for i in self.items: self.tree.insert('', tk.END, values=(i['id'], i['name']))

    def _filter_items(self, *args):
        st = self.search_var.get().lower()
        for i in self.tree.get_children(): self.tree.delete(i)
        for i in self.items:
            if st in str(i['name']).lower() or st in str(i['id']).lower(): self.tree.insert('', tk.END, values=(i['id'], i['name']))

    def _on_item_select(self, event):
        sel = self.tree.selection()
        if not sel: return
        iid = int(self.tree.item(sel[0])['values'][0])
        for i in self.items:
            if i['id'] == iid: self.current_row = i; self._show_editor(i); return

    def _show_editor(self, item):
        for w in self.right_container.winfo_children(): w.destroy()
        
        # 顶部：物品信息(去掉原始备注区域)
        top_frame = tk.Frame(self.right_container)
        top_frame.pack(fill=tk.X, pady=(0, 3))
        
        tk.Label(top_frame, text=f'当前编辑: {item["name"]} (ID: {item["id"]})', 
                font=('微软雅黑', 11, 'bold'), fg='#2196F3').pack(anchor=tk.W)

        # 中间：编辑区和预览区分割窗口(加高)
        split = tk.PanedWindow(self.right_container, orient=tk.HORIZONTAL)
        split.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)

        # 左侧：可视化预览编辑区
        pfr = tk.Frame(split)
        split.add(pfr, width=400)
        
        tk.Label(pfr, text='可视化预览 (可编辑):', font=('微软雅黑', 9, 'bold'), bg='#e3f2fd').pack(fill=tk.X)
        self.preview = tk.Text(pfr, wrap=tk.WORD, font=('微软雅黑', 10), bg='#ffffff', 
                              undo=True, selectbackground='#b3d4fc', selectforeground='black',
                              inactiveselectbackground='#d0e8ff')
        ps = tk.Scrollbar(pfr, command=self.preview.yview)
        self.preview.config(yscrollcommand=ps.set)
        self.preview.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        ps.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 绑定预览区的修改事件
        self.preview.bind('<KeyRelease>', lambda e: self._on_preview_key())
        self.preview.bind('<ButtonRelease>', lambda e: self._on_preview_selection_change())

        # 中间操作按钮区域（垂直居中）
        action_frame = tk.Frame(split, bg='#e0e0e0', width=60)
        split.add(action_frame)
        action_frame.pack_propagate(False)
        
        # 使用中间容器实现垂直居中
        btn_container = tk.Frame(action_frame, bg='#e0e0e0')
        btn_container.place(relx=0.5, rely=0.5, anchor=tk.CENTER)
        
        # 从代码区同步到预览区
        tk.Button(btn_container, text='←预 览', command=self._update_preview, 
                 font=('微软雅黑', 10, 'bold'), width=8, bg='#4CAF50', fg='white',
                 relief=tk.RAISED, cursor='hand2').pack(pady=5)
        
        # 从预览区同步到代码区
        tk.Button(btn_container, text='代 码→', command=self._sync_from_preview, 
                 font=('微软雅黑', 10, 'bold'), width=8, bg='#2196F3', fg='white',
                 relief=tk.RAISED, cursor='hand2').pack(pady=5)

        # 右侧：源码编辑区
        efr = tk.Frame(split)
        split.add(efr, width=400)
        tk.Label(efr, text='编辑源码:', font=('微软雅黑', 9, 'bold'), bg='#e3f2fd').pack(fill=tk.X)
        self.editor = tk.Text(efr, wrap=tk.WORD, font=('Consolas', 10), undo=True,
                             selectbackground='#b3d4fc', selectforeground='black',
                             inactiveselectbackground='#d0e8ff')
        self.editor.insert('1.0', item['desc'].replace('[br]', '\n').replace('<br>', '\n').replace('<br/>', '\n'))
        es = tk.Scrollbar(efr, command=self.editor.yview)
        self.editor.config(yscrollcommand=es.set)
        self.editor.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        es.pack(side=tk.RIGHT, fill=tk.Y)
        # 绑定编辑区的修改事件（回车不同步到预览区）
        self.editor.bind('<KeyRelease>', lambda e: None)

        # 底部：工具栏区域
        tool_frame = tk.Frame(self.right_container, bg='#f0f0f0')
        tool_frame.pack(fill=tk.X, padx=5, pady=3)
        
        # 第一行：编辑模式 + 格式化工具
        format_frame = tk.Frame(tool_frame, bg='#f0f0f0')
        format_frame.pack(fill=tk.X, pady=(3, 3))

        tk.Label(format_frame, text='模式:', font=('微软雅黑', 9, 'bold'), bg='#f0f0f0').pack(side=tk.LEFT, padx=3)
        self.edit_mode_var = tk.StringVar(value='UBB')
        ttk.Radiobutton(format_frame, text='UBB', variable=self.edit_mode_var, value='UBB',
                       command=self._on_mode_change).pack(side=tk.LEFT, padx=3)
        ttk.Radiobutton(format_frame, text='HTML', variable=self.edit_mode_var, value='HTML',
                       command=self._on_mode_change).pack(side=tk.LEFT, padx=3)

        tk.Label(format_frame, text='|', font=('微软雅黑', 9), bg='#f0f0f0').pack(side=tk.LEFT, padx=5)

        tk.Label(format_frame, text='格式:', font=('微软雅黑', 9, 'bold'), bg='#f0f0f0').pack(side=tk.LEFT, padx=3)
        tk.Button(format_frame, text='B 粗体', command=lambda: self._apply_editor_format('bold'), 
                 font=('微软雅黑', 9, 'bold'), width=7).pack(side=tk.LEFT, padx=2)
        tk.Button(format_frame, text='I 斜体', command=lambda: self._apply_editor_format('italic'), 
                 font=('微软雅黑', 9, 'italic'), width=7).pack(side=tk.LEFT, padx=2)
        tk.Button(format_frame, text='U 下划线', command=lambda: self._apply_editor_format('underline'), 
                 font=('微软雅黑', 9), width=9).pack(side=tk.LEFT, padx=2)
        tk.Button(format_frame, text='清除格式', command=self._clear_all_format, 
                 width=9, bg='#FF5722', fg='white', font=('微软雅黑', 9)).pack(side=tk.LEFT, padx=5)

        # 第二行：字号和颜色
        size_color_frame = tk.Frame(tool_frame, bg='#f0f0f0')
        size_color_frame.pack(fill=tk.X, pady=(3, 3))

        tk.Label(size_color_frame, text='字号:', font=('微软雅黑', 9), bg='#f0f0f0').pack(side=tk.LEFT, padx=3)
        self.size_var = tk.StringVar(value='12')
        ttk.Combobox(size_color_frame, textvariable=self.size_var, values=[str(i) for i in range(10, 81, 2)],
                    width=5, state='readonly', font=('微软雅黑', 9)).pack(side=tk.LEFT, padx=2)
        tk.Button(size_color_frame, text='应用字号', command=self._apply_size, width=9, font=('微软雅黑', 9)).pack(side=tk.LEFT, padx=2)

        tk.Label(size_color_frame, text='|', font=('微软雅黑', 9), bg='#f0f0f0').pack(side=tk.LEFT, padx=5)

        tk.Button(size_color_frame, text='选择颜色', command=self._apply_custom_color, 
                 width=10, bg='#FFEB3B', font=('微软雅黑', 9)).pack(side=tk.LEFT, padx=2)

        # 第三行：格式说明
        hint_frame = tk.Frame(tool_frame, bg='#f0f0f0')
        hint_frame.pack(fill=tk.X, pady=(3, 3))

        self.hint_label = tk.Label(hint_frame, text='', 
                                   font=('微软雅黑', 8), fg='#666', bg='#f0f0f0', justify=tk.LEFT)
        self.hint_label.pack(anchor=tk.W, padx=5)

        self._update_preview()
        self._update_mode_hint()

    def _schedule_preview(self):
        if self.preview_timer: self.root.after_cancel(self.preview_timer)
        self.preview_timer = self.root.after(300, self._update_preview)

    def _on_mode_change(self):
        """编辑模式切换时的处理"""
        self._update_mode_hint()

    def _update_mode_hint(self):
        """更新格式提示文字"""
        if self.edit_mode_var.get() == 'HTML':
            self.hint_label.config(text='HTML 格式：<font size=50>大字</font> <font color=red>红色</font> <br>换行 <b>粗体</b> <i>斜体</i> <u>下划线</u>')
        else:
            self.hint_label.config(text='UBB 格式：[size=50]大字[/size] [color=red]红色[/color] [br]换行 [b]粗体[/b] [i]斜体[/i] [u]下划线[/u]')

    def _apply_editor_format(self, fmt_type):
        """在编辑区或预览区应用格式（粗体/斜体/下划线），并同步"""
        # 优先检查预览区是否有选中
        try:
            ps, pe = self.preview.index(tk.SEL_FIRST), self.preview.index(tk.SEL_LAST)
            # 预览区有选中，直接在预览区应用tag，然后同步到代码
            selected_text = self.preview.get(ps, pe)
            
            tag_name = f'fmt_{fmt_type}'
            current_tags = self.preview.tag_names(ps)
            if tag_name in current_tags:
                self.preview.tag_remove(tag_name, ps, pe)
            else:
                if fmt_type == 'bold':
                    self.preview.tag_add(tag_name, ps, pe)
                    self.preview.tag_configure(tag_name, font=('微软雅黑', 10, 'bold'))
                elif fmt_type == 'italic':
                    self.preview.tag_add(tag_name, ps, pe)
                    self.preview.tag_configure(tag_name, font=('微软雅黑', 10, 'italic'))
                elif fmt_type == 'underline':
                    self.preview.tag_add(tag_name, ps, pe)
                    self.preview.tag_configure(tag_name, underline=True)
            
            # 同步到代码区
            self._sync_from_preview()
            return
        except tk.TclError:
            pass  # 预览区没有选中，继续检查编辑区
        
        # 编辑区有选中
        try:
            s, e = self.editor.index(tk.SEL_FIRST), self.editor.index(tk.SEL_LAST)
            raw = self.editor.get(s, e)
            
            mode = self.edit_mode_var.get()
            
            # 检测是否已有相同标签
            if mode == 'HTML':
                # 检查是否已有相同标签
                if fmt_type == 'bold' and raw.strip().startswith('<b>') and raw.strip().endswith('</b>'):
                    # 已有粗体标签，移除它
                    clean = raw.strip()[3:-4]
                elif fmt_type == 'italic' and raw.strip().startswith('<i>') and raw.strip().endswith('</i>'):
                    clean = raw.strip()[3:-4]
                elif fmt_type == 'underline' and raw.strip().startswith('<u>') and raw.strip().endswith('</u>'):
                    clean = raw.strip()[3:-4]
                else:
                    # 没有相同标签，添加新标签
                    clean = strip_html_tags(raw)
                    if fmt_type == 'bold':
                        clean = f'<b>{clean}</b>'
                    elif fmt_type == 'italic':
                        clean = f'<i>{clean}</i>'
                    elif fmt_type == 'underline':
                        clean = f'<u>{clean}</u>'
            else:  # UBB模式
                if fmt_type == 'bold' and raw.strip().startswith('[b]') and raw.strip().endswith('[/b]'):
                    clean = raw.strip()[3:-4]
                elif fmt_type == 'italic' and raw.strip().startswith('[i]') and raw.strip().endswith('[/i]'):
                    clean = raw.strip()[3:-4]
                elif fmt_type == 'underline' and raw.strip().startswith('[u]') and raw.strip().endswith('[/u]'):
                    clean = raw.strip()[3:-4]
                else:
                    clean = strip_ubb_tags(raw)
                    if fmt_type == 'bold':
                        clean = f'[b]{clean}[/b]'
                    elif fmt_type == 'italic':
                        clean = f'[i]{clean}[/i]'
                    elif fmt_type == 'underline':
                        clean = f'[u]{clean}[/u]'
            
            self.editor.delete(s, e)
            self.editor.insert(s, clean)
            # 自动同步到预览区
            self._update_preview()
            # 清除选中
            self.editor.tag_remove(tk.SEL, '1.0', tk.END)
        except tk.TclError:
            messagebox.showwarning('提示', '请先选择文字')

    def _clear_all_format(self):
        """清除所有格式，保留换行标签"""
        try:
            # 获取编辑区所有内容
            content = self.editor.get('1.0', tk.END).strip()
            
            # 移除所有格式标签，保留[br]或<br>
            mode = self.edit_mode_var.get()
            if mode == 'HTML':
                # 移除所有HTML标签，保留<br>
                cleaned = re.sub(r'<(?!br\s*/?>)[^>]*>', '', content)
                # 清理多余空格和空行
                lines = cleaned.split('\n')
                cleaned = '\n'.join([line.strip() for line in lines if line.strip()])
            else:
                # 移除所有UBB标签，保留[br]
                cleaned = re.sub(r'\[(?!br)[^\]]*\]', '', content)
                # 清理多余空格和空行
                lines = cleaned.split('\n')
                cleaned = '\n'.join([line.strip() for line in lines if line.strip()])
            
            # 更新编辑区
            self.editor.delete('1.0', tk.END)
            self.editor.insert('1.0', cleaned)
            
            # 清除预览区所有自定义tag
            for tag in self.preview.tag_names():
                if tag.startswith('fmt_') or tag.startswith('u_') or tag.startswith('h_'):
                    self.preview.tag_delete(tag)
            
            # 同步到预览区
            self._update_preview()
        except Exception as e:
            messagebox.showerror('错误', str(e))

    def _sync_to_preview(self):
        """从编辑区同步到预览区"""
        self._update_preview()
        messagebox.showinfo('成功', '已从编辑区同步到预览区')

    def _on_preview_key(self, event=None):
        """预览区按键事件处理,回车键时在代码区生成br标签"""
        if event and event.keysym == 'Return':
            # 延迟同步到代码区,让回车符先插入到预览区
            self.root.after(50, self._sync_from_preview)

    def _schedule_sync_to_preview(self):
        """定时从预览区同步到编辑区（已废弃）"""
        pass

    def _on_preview_selection_change(self):
        """预览区选中变化时的处理"""
        pass  # 可以在这里添加选中状态变化的逻辑

    def _apply_preview_format(self, fmt_type):
        """在预览区应用格式（粗体/斜体/下划线）"""
        try:
            s, e = self.preview.index(tk.SEL_FIRST), self.preview.index(tk.SEL_LAST)
            selected_text = self.preview.get(s, e)
            
            # 应用tag
            tag_name = f'fmt_{fmt_type}'
            if fmt_type == 'bold':
                current_tags = self.preview.tag_names(s)
                if tag_name in current_tags:
                    self.preview.tag_remove(tag_name, s, e)
                else:
                    self.preview.tag_add(tag_name, s, e)
                    self.preview.tag_configure(tag_name, font=('微软雅黑', 10, 'bold'))
            elif fmt_type == 'italic':
                current_tags = self.preview.tag_names(s)
                if tag_name in current_tags:
                    self.preview.tag_remove(tag_name, s, e)
                else:
                    self.preview.tag_add(tag_name, s, e)
                    self.preview.tag_configure(tag_name, font=('微软雅黑', 10, 'italic'))
            elif fmt_type == 'underline':
                current_tags = self.preview.tag_names(s)
                if tag_name in current_tags:
                    self.preview.tag_remove(tag_name, s, e)
                else:
                    self.preview.tag_add(tag_name, s, e)
                    self.preview.tag_configure(tag_name, underline=True)
        except tk.TclError:
            messagebox.showwarning('提示', '请先在预览区选择文字')

    def _apply_preview_size(self):
        """在预览区应用字号"""
        try:
            s, e = self.preview.index(tk.SEL_FIRST), self.preview.index(tk.SEL_LAST)
            size = int(self.preview_size_var.get())
            
            tag_name = f'fmt_size_{size}'
            self.preview.tag_add(tag_name, s, e)
            self.preview.tag_configure(tag_name, font=('微软雅黑', size))
        except tk.TclError:
            messagebox.showwarning('提示', '请先在预览区选择文字')

    def _apply_preview_color(self):
        """在预览区应用颜色"""
        try:
            s, e = self.preview.index(tk.SEL_FIRST), self.preview.index(tk.SEL_LAST)
            c = colorchooser.askcolor(title='选择颜色')
            if not c[1]: return
            
            tag_name = f'fmt_color_{c[1].upper()}'
            self.preview.tag_add(tag_name, s, e)
            self.preview.tag_configure(tag_name, foreground=c[1])
        except tk.TclError:
            messagebox.showwarning('提示', '请先在预览区选择文字')

    def _sync_from_preview(self):
        """从预览区反向生成代码到编辑区"""
        if not hasattr(self, 'preview'): return

        # 获取预览区所有内容
        content = self.preview.get('1.0', tk.END)

        # 使用当前编辑模式生成对应的UBB/HTML代码
        mode = self.edit_mode_var.get()
        result = self._generate_code_from_preview(mode)

        # 更新编辑区
        self.editor.delete('1.0', tk.END)
        self.editor.insert('1.0', result)

    def _generate_code_from_preview(self, mode):
        """从预览区的tag生成对应的UBB/HTML代码"""
        content = self.preview.get('1.0', tk.END)

        # 获取所有tag的范围
        tags_info = self._get_all_tags_with_ranges()

        # 构建结果
        result = self._apply_tags_to_text(content, tags_info, mode.lower())
        return result

    def _get_all_tags_with_ranges(self):
        """获取所有tag及其作用范围"""
        tags_info = {}
        line = 1
        while True:
            try:
                line_content = self.preview.get(f'{line}.0', f'{line}.end')
                if not line_content and line > 1:
                    break
                for col in range(len(line_content)):
                    pos = f'{line}.{col}'
                    tags = self.preview.tag_names(pos)
                    for tag in tags:
                        if tag.startswith('fmt_') or tag.startswith('u_') or tag.startswith('h_'):
                            if tag not in tags_info:
                                tags_info[tag] = []
                            tags_info[tag].append(pos)
                line += 1
            except:
                break
        return tags_info

    def _apply_tags_to_text(self, text, tags_info, fmt):
        """将tag信息应用到文本，生成对应的代码"""
        # 简化实现：逐字符检查tag并生成代码
        lines = text.split('\n')
        result_lines = []

        for line_idx, line in enumerate(lines):
            # 处理换行符 - 在行间添加br标签
            if line_idx > 0 and line.strip():
                # 如果不是第一行且有内容,添加br标签
                br_tag = '<br>' if fmt == 'html' else '[br]'
                result_lines.append(br_tag)
            
            line_result = ''
            i = 0
            while i < len(line):
                pos = f'{line_idx + 1}.{i}'
                tags = self.preview.tag_names(pos)

                # 获取当前字符的格式
                char_fmt = self._parse_tags(tags)

                # 查找连续相同格式的字符
                j = i + 1
                while j < len(line):
                    next_pos = f'{line_idx + 1}.{j}'
                    next_tags = self.preview.tag_names(next_pos)
                    next_fmt = self._parse_tags(next_tags)
                    if next_fmt != char_fmt:
                        break
                    j += 1

                # 生成代码（去掉多余空格）
                segment = line[i:j].strip()
                if segment:
                    if char_fmt:
                        if fmt == 'html':
                            line_result += self._format_to_html(char_fmt, segment)
                        else:
                            line_result += self._format_to_ubb(char_fmt, segment)
                    else:
                        line_result += segment

                i = j

            if line_result.strip():
                result_lines.append(line_result)

        return '\n'.join(result_lines)

    def _parse_tags(self, tags):
        """解析tag列表为格式信息"""
        fmt = {'bold': False, 'italic': False, 'underline': False, 'size': 10, 'color': 'black'}
        for tag in tags:
            if tag == 'fmt_bold':
                fmt['bold'] = True
            elif tag == 'fmt_italic':
                fmt['italic'] = True
            elif tag == 'fmt_underline':
                fmt['underline'] = True
            elif tag.startswith('fmt_size_'):
                fmt['size'] = int(tag.split('_')[-1])
            elif tag.startswith('fmt_color_'):
                fmt['color'] = tag.replace('fmt_color_', '')
            elif tag.startswith('u_') or tag.startswith('h_'):
                # 解析原有的UBB/HTML tag
                parts = tag.split('_')
                if len(parts) >= 5:
                    fmt['size'] = int(parts[1])
                    fmt['color'] = parts[2]
                    if parts[3] == 'bold':
                        fmt['bold'] = True
                    if parts[4] == 'italic':
                        fmt['italic'] = True
                    if len(parts) > 5 and parts[5] == 'True':
                        fmt['underline'] = True
        return fmt

    def _format_to_ubb(self, fmt, text):
        """将格式转换为UBB代码"""
        result = text
        if fmt['color'] != 'black':
            result = f'[color={fmt["color"]}]{result}[/color]'
        if fmt['size'] != 10:
            result = f'[size={fmt["size"]}]{result}[/size]'
        if fmt['underline']:
            result = f'[u]{result}[/u]'
        if fmt['italic']:
            result = f'[i]{result}[/i]'
        if fmt['bold']:
            result = f'[b]{result}[/b]'
        return result

    def _format_to_html(self, fmt, text):
        """将格式转换为HTML代码"""
        result = text
        if fmt['underline']:
            result = f'<u>{result}</u>'
        if fmt['italic']:
            result = f'<i>{result}</i>'
        if fmt['bold']:
            result = f'<b>{result}</b>'
        if fmt['size'] != 10 or fmt['color'] != 'black':
            attrs = ''
            if fmt['size'] != 10:
                attrs += f' size={fmt["size"]}'
            if fmt['color'] != 'black':
                attrs += f' color={fmt["color"]}'
            result = f'<font{attrs}>{result}</font>'
        return result

    def _update_preview(self):
        if not hasattr(self, 'preview'): return
        raw_text = self.editor.get('1.0', tk.END)

        # 保存当前光标位置和选中范围
        try:
            cursor_pos = self.preview.index(tk.INSERT)
            sel_start = self.preview.index(tk.SEL_FIRST)
            sel_end = self.preview.index(tk.SEL_LAST)
        except:
            cursor_pos = '1.0'
            sel_start = None
            sel_end = None

        self.preview.delete('1.0', tk.END)
        # 清除所有tag
        for t in self.preview.tag_names():
            self.preview.tag_delete(t)

        # 检测格式
        fmt = detect_format(raw_text)
        
        if fmt == 'html':
            self._render_html_preview(raw_text)
        else:
            self._render_ubb_preview(raw_text)

        # 恢复光标和选中范围
        try:
            self.preview.mark_set(tk.INSERT, cursor_pos)
            if sel_start and sel_end:
                self.preview.tag_remove(tk.SEL, '1.0', tk.END)
                self.preview.tag_add(tk.SEL, sel_start, sel_end)
        except:
            pass

    def _render_ubb_preview(self, raw_text):
        """渲染 UBB 格式的预览"""
        pat = re.compile(r'(\[/?[a-zA-Z]+[^]]*\])', re.IGNORECASE)
        parts = pat.split(raw_text)

        sz = 10
        cl = 'black'
        is_bold = False
        is_italic = False
        is_underline = False

        for p in parts:
            if not p: continue
            lp = p.lower()

            if lp.startswith('[size='):
                m = re.search(r'=(\d+)', p)
                if m: sz = int(m.group(1))
            elif lp == '[/size]':
                sz = 10
            elif lp.startswith('[color='):
                m = re.search(r'=([^]]+)', p)
                if m: cl = m.group(1)
            elif lp == '[/color]':
                cl = 'black'
            elif lp == '[b]':
                is_bold = True
            elif lp == '[/b]':
                is_bold = False
            elif lp == '[i]':
                is_italic = True
            elif lp == '[/i]':
                is_italic = False
            elif lp == '[u]':
                is_underline = True
            elif lp == '[/u]':
                is_underline = False
            elif lp == '[br]':
                s = self.preview.index(tk.INSERT)
                self.preview.insert(tk.END, '\n')
            else:
                # 去除多余空格和换行
                clean_p = p.strip()
                if clean_p:
                    s = self.preview.index(tk.INSERT)
                    self.preview.insert(tk.END, clean_p)
                    e = self.preview.index(tk.INSERT)

                    font_weight = 'bold' if is_bold else 'normal'
                    font_slant = 'italic' if is_italic else 'roman'
                    font_underline = True if is_underline else False

                    tag_name = f'u_{sz}_{cl}_{font_weight}_{font_slant}_{font_underline}'
                    font = ('微软雅黑', sz, font_weight, font_slant)
                    self.preview.tag_configure(tag_name, font=font, foreground=cl, underline=font_underline)
                    self.preview.tag_add(tag_name, s, e)

    def _render_html_preview(self, raw_text):
        """渲染 HTML 格式的预览"""
        # 保存原始文本用于解析
        temp_text = raw_text
        
        # 默认值
        default_size = 10
        current_size = default_size
        current_color = 'black'
        is_bold = False
        is_italic = False
        is_underline = False
        
        # 合并UBB和HTML标签的正则
        combined_pattern = re.compile(
            r'(<font[^>]*>|</font>|<b>|</b>|<i>|</i>|<u>|</u>|<br\s*/?>|<br>|'
            r'\[size=[^\]]*\]|\[/size\]|\[color=[^\]]*\]|\[/color\]|\[b\]|\[/b\]|\[i\]|\[/i\]|\[u\]|\[/u\]|\[br\])',
            re.IGNORECASE
        )
        
        parts = combined_pattern.split(temp_text)
        
        for p in parts:
            if not p:
                continue
            lp = p.lower()
            
            # HTML 字体标签
            if lp.startswith('<font'):
                size_match = re.search(r'size=(\d+)', p)
                color_match = re.search(r'color=([^>\s]+)', p)
                if size_match:
                    current_size = int(size_match.group(1))
                if color_match:
                    current_color = color_match.group(1).strip('"').strip("'")
            elif lp == '</font>':
                current_size = default_size
                current_color = 'black'
            # HTML 粗体
            elif lp == '<b>':
                is_bold = True
            elif lp == '</b>':
                is_bold = False
            # HTML 斜体
            elif lp == '<i>':
                is_italic = True
            elif lp == '</i>':
                is_italic = False
            # HTML 下划线
            elif lp == '<u>':
                is_underline = True
            elif lp == '</u>':
                is_underline = False
            # HTML 换行
            elif lp.startswith('<br'):
                s = self.preview.index(tk.INSERT)
                self.preview.insert(tk.END, '\n')
            # UBB 标签兼容处理
            elif lp.startswith('[size='):
                m = re.search(r'=(\d+)', p)
                if m: current_size = int(m.group(1))
            elif lp == '[/size]':
                current_size = default_size
            elif lp.startswith('[color='):
                m = re.search(r'=([^]]+)', p)
                if m: current_color = m.group(1)
            elif lp == '[/color]':
                current_color = 'black'
            elif lp == '[b]':
                is_bold = True
            elif lp == '[/b]':
                is_bold = False
            elif lp == '[i]':
                is_italic = True
            elif lp == '[/i]':
                is_italic = False
            elif lp == '[u]':
                is_underline = True
            elif lp == '[/u]':
                is_underline = False
            elif lp == '[br]':
                s = self.preview.index(tk.INSERT)
                self.preview.insert(tk.END, '\n')
            else:
                # 普通文本 - 去除多余空格和换行
                clean_p = p.strip()
                if clean_p:
                    s = self.preview.index(tk.INSERT)
                    self.preview.insert(tk.END, clean_p)
                    e = self.preview.index(tk.INSERT)

                    font_weight = 'bold' if is_bold else 'normal'
                    font_slant = 'italic' if is_italic else 'roman'
                    font_underline = True if is_underline else False

                    tag_name = f'h_{current_size}_{current_color}_{font_weight}_{font_slant}_{font_underline}'
                    font = ('微软雅黑', current_size, font_weight, font_slant)
                    self.preview.tag_configure(tag_name, font=font, foreground=current_color, underline=font_underline)
                    self.preview.tag_add(tag_name, s, e)

    def _apply_size(self, event=None):
        """应用字号，支持预览区或编辑区选中，并同步"""
        sz = self.size_var.get()
        if sz == '12': sz = '12'  # 默认值
        
        # 优先检查预览区是否有选中
        try:
            ps, pe = self.preview.index(tk.SEL_FIRST), self.preview.index(tk.SEL_LAST)
            size = int(sz)
            tag_name = f'fmt_size_{size}'
            self.preview.tag_add(tag_name, ps, pe)
            self.preview.tag_configure(tag_name, font=('微软雅黑', size))
            # 同步到代码区
            self._sync_from_preview()
            return
        except tk.TclError:
            pass  # 预览区没有选中，继续检查编辑区
        
        # 编辑区有选中
        try:
            s, e = self.editor.index(tk.SEL_FIRST), self.editor.index(tk.SEL_LAST)
            raw = self.editor.get(s, e)
            clean = strip_ubb_tags(strip_html_tags(raw))
            
            # 使用当前编辑模式
            mode = self.edit_mode_var.get()
            if mode == 'HTML':
                new = f'<font size={sz}>{clean}</font>'
            else:
                new = f'[size={sz}]{clean}[/size]'
            
            self.editor.delete(s, e)
            self.editor.insert(s, new)
            # 自动同步到预览区
            self._update_preview()
            # 清除选中
            self.editor.tag_remove(tk.SEL, '1.0', tk.END)
        except tk.TclError:
            messagebox.showwarning('提示', '请先选择文字')

    def _apply_custom_color(self):
        """应用颜色，支持预览区或编辑区选中，并同步"""
        c = colorchooser.askcolor(title='选择颜色')
        if not c[1]: return
        
        # 优先检查预览区是否有选中
        try:
            ps, pe = self.preview.index(tk.SEL_FIRST), self.preview.index(tk.SEL_LAST)
            tag_name = f'fmt_color_{c[1].upper()}'
            self.preview.tag_add(tag_name, ps, pe)
            self.preview.tag_configure(tag_name, foreground=c[1])
            # 同步到代码区
            self._sync_from_preview()
            return
        except tk.TclError:
            pass  # 预览区没有选中，继续检查编辑区
        
        # 编辑区有选中
        try:
            s, e = self.editor.index(tk.SEL_FIRST), self.editor.index(tk.SEL_LAST)
            raw = self.editor.get(s, e)
            clean = strip_ubb_tags(strip_html_tags(raw))
            hex_color = c[1].upper()
            
            # 使用当前编辑模式
            mode = self.edit_mode_var.get()
            if mode == 'HTML':
                new = f'<font color={hex_color}>{clean}</font>'
            else:
                new = f'[color={hex_color}]{clean}[/color]'
            
            self.editor.delete(s, e)
            self.editor.insert(s, new)
            # 自动同步到预览区
            self._update_preview()
            # 清除选中
            self.editor.tag_remove(tk.SEL, '1.0', tk.END)
        except tk.TclError:
            messagebox.showwarning('提示', '请先选择文字')

    def _save_desc(self):
        if not self.current_row or not self.file_path: return messagebox.showwarning('警告', '请先选择并打开文件')
        try:
            content = self.editor.get('1.0', tk.END).strip()
            
            # 根据编辑模式处理换行符
            mode = self.edit_mode_var.get()
            if mode == 'HTML':
                new_desc = content.replace('\r\n', '<br>').replace('\n', '<br>')
            else:
                new_desc = content.replace('\r\n', '[br]').replace('\n', '[br]')
            
            row = self.current_row['row']

            if self.file_type == '.xls':
                rb = xlrd.open_workbook(self.file_path, formatting_info=False)
                wb = xl_copy(rb)
                wb.get_sheet(0).write(row, 25, new_desc)
                tmp = self.file_path + '.tmp'
                wb.save(tmp)
                os.replace(tmp, self.file_path)
                self.xlrd_book = xlrd.open_workbook(self.file_path, formatting_info=False)
            else:
                self.wb.active.cell(row=row, column=26).value = new_desc
                self.wb.save(self.file_path)

            self.current_row['desc'] = new_desc
            messagebox.showinfo('成功', f'已保存回原文件 ({mode}模式)')
        except Exception as e: messagebox.showerror('错误', str(e))

    def _show_log(self):
        w = tk.Toplevel(self.root)
        w.title('日志')
        w.geometry('800x500')
        t = tk.Text(w, wrap=tk.WORD)
        t.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        t.insert('1.0', self.logger.get_log_content())
        t.config(state=tk.DISABLED)
        f = tk.Frame(w)
        f.pack(fill=tk.X, padx=5, pady=5)
        tk.Button(f, text='刷新', command=lambda: (t.config(state=tk.NORMAL), t.delete('1.0',tk.END), t.insert('1.0',self.logger.get_log_content()), t.config(state=tk.DISABLED))).pack(side=tk.LEFT, padx=5)
        tk.Button(f, text='关闭', command=w.destroy).pack(side=tk.RIGHT, padx=5)

    def run(self):
        self.root.mainloop()

if __name__ == '__main__':
    ItemNoteEditor().run()
